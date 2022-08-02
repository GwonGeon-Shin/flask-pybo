#!/usr/bin/env python
# coding: utf-8

# In[ ]:


#from crypt import methods
#from crypt import methods
from datetime import datetime
from re import search

from flask import Blueprint, render_template, request, url_for, session, g
from werkzeug.utils import redirect, secure_filename

from pybo import db
from pybo.models import Question, User
from ..forms import AnswerForm

from pybo.views.auth_views import login_required

from openpyxl import load_workbook

bp = Blueprint('question', __name__, url_prefix='/question')

@bp.route('/list')
@login_required
def _list():
    page = request.args.get('page', type=int, default=1) #페이지
    kw = request.args.get('kw', type=str, default='')


    # ▼▼▼ 정렬 기능은 조금더 공부한 다음에 적용 하자. 전부 주석화
    # 정렬 기능 부분 시작
    # so = request.args.get('so', type=str, default='recent')
    # if so == 'recent':
    #     question_list = Question.query.order_by(Question.date) #출원일 최신 기준
    # elif so == 'recommend':
    #     question_list = Question.query.order_by(Question.p_name)
    # else:
    #     question_list = Question.query.order_by(Question.create_date) #엑셀 데이터 최상단을 위로 하기위해서 .decs()제거
    # 정렬 기능 부분 종료
    # ▲▲▲ 정렬 기능은 조금더 공부한 다음에 적용 하자.

    #question_list=Question.query.order_by(Question.create_date)

    ### ▼▼▼ 사용자별 question_list 구별해서 구현하기 시작

    if g.user:
        question_list=Question.query.order_by(Question.create_date)
        if g.user.username == "관리자":
            #question_list_01 = question_list \
            #    .filter(Question.req_name.ilike('%권다현%')
            #    ) # 의뢰인 자료 기준 검색 req_name)
            #question_list = question_list_01.paginate(page, per_page=10)


            search = '%%{}%%' .format(kw)
            question_list = question_list \
                .filter(Question.p_name.ilike(search) | # 질문 제목 subject => 국문 명칭자료 p_name
                        Question.req_name.ilike(search) | # 의뢰인 자료 기준 검색 req_name
                        Question.our_ref.ilike(search) # 관리번호 기준 검색 our_ref
                        )   


            question_list=question_list.paginate(page, per_page=10)
            return render_template('question/question_list.html', question_list=question_list, page=page, kw=kw)
        else:
            name_test_100 = g.user.username # User.username 가 아닌 g.user.username로 했다.
            #User.username 으로 할시 로그아웃 하고 다른 이름으로 로그인할때 기존 User.username가 남아있을수 있어서.
            name_test_101 = "%"+name_test_100+"%"
            question_list_02 = question_list \
                .filter(Question.req_name.ilike(name_test_101))


            search = '%%{}%%' .format(kw)
            question_list_02 = question_list_02 \
                .filter(Question.p_name.ilike(search) | # 질문 제목 subject => 국문 명칭자료 p_name
                        Question.req_name.ilike(search) | # 의뢰인 자료 기준 검색 req_name
                        Question.our_ref.ilike(search) # 관리번호 기준 검색 our_ref
                        )

            question_list = question_list_02.paginate(page, per_page=10)
            return render_template('question/question_list.html', question_list=question_list, page=page, kw=kw)
            #pass
    else:
        question_list=Question.query.order_by(Question.create_date)
        #question_list=""


    ### ▲▲▲ 사용자별 question_list 구별해서 구현하기 종료


    #question_list=Question.query.order_by(Question.create_date) #엑셀 데이터 최상단을 위로 하기위해서 .decs()제거
    # 책 예제 기준, 기존 question_list 표현 방식임
    # ㄴ> 정렬기능, 사용자별 리스트 표시 기능 위해서 해당 구문 주석화

    #question_list=Question.query.order_by(Question.create_date.desc())
    # 이하 검색을 위한 기능 구현
    # 기존 예제는 outerjoin 이였지만 우리는 Quesiton 모델 하나만 쓰기에
    # join 형식으로 검색 구현을 해본다.
    if kw:
        search = '%%{}%%' .format(kw)
        question_list = question_list \
            .filter(Question.p_name.ilike(search) | # 질문 제목 subject => 국문 명칭자료 p_name
                    Question.req_name.ilike(search) | # 의뢰인 자료 기준 검색 req_name
                    Question.our_ref.ilike(search) # 관리번호 기준 검색 our_ref
                    )        
    question_list = question_list.paginate(page, per_page=10)
    return render_template('question/question_list.html', question_list=question_list, page=page, kw=kw)

@bp.route('/userlist')
@login_required
def userlist():
    user_list = User.query.all()
    return render_template('question/user_list.html', user_list=user_list)


@bp.route('/detail/<int:question_id>/')
@login_required
def detail(question_id):
    form = AnswerForm()
    question = Question.query.get_or_404(question_id)
    return render_template('question/question_detail.html', question=question, form=form)
    
@bp.route('/create/<int:question_id>', methods=("POST",))
@login_required
def create(question_id):
    form = AnswerForm()
    question = Question.query.get_or_404(question_id)

    if form.validate_on_submit():
        pay_date = request.form['content']
        question.pay_date = pay_date
        #question.create_date = datetime.now()
        #answer=Question(pay_date=pay_date, create_date=datetime.now())
        #question.append(answer)
        db.session.commit()
        return redirect(url_for('question.detail', question_id=question_id))
    return render_template('question/question_detail.html', question=question, form=form)

@bp.route('/upload_file/')
@login_required
def upload_file():
    return render_template('question/upload_file.html')

@bp.route('/uploader', methods=['GET', 'POST'])
@login_required
def uploader():
    if request.method =='POST':
        f=request.files['file']
        f.save(secure_filename(f.filename))

        #데이터 삭제 부분 시작
        query_count_num = Question.query.count()
        for query_del_num in range(1,query_count_num+1):
            query_del_num_01 = Question.query.get(query_del_num)
            db.session.delete(query_del_num_01)
            db.session.commit()
            db.session.close()

        # 데이터 입력 부분 시작
        wb = load_workbook(filename='db_test.xlsx')
        ws = wb.active
        for row in range(1, ws.max_row+1):
            try:
                question = Question(our_ref=ws.cell(row,1).value, pay_date=ws.cell(row,2).value, req_name=ws.cell(row,3).value, cus_name=ws.cell(row,4).value, date=ws.cell(row,5).value, number=ws.cell(row,6).value, q_date=ws.cell(row,7).value, q_d_date=ws.cell(row,8).value, d_d_date=ws.cell(row,9).value, d_d_num=ws.cell(row,10).value, p_name=ws.cell(row,11).value, charge_01=ws.cell(row,12).value, charge_02=ws.cell(row,13).value, p_charge=ws.cell(row,14).value, create_date=datetime.now())
                db.session.add(question)
                db.session.commit()
            except:
                db.session.rollback()
                pass
            finally:
                db.session.close()
        return redirect(url_for('question._list'))
        #return render_template('question/question_list.html')
        #return '파일이 저장 되었습니다.' + str(Question.query.count())
        #return render_template('question/upload_file.html')
    return render_template('question/question_list.html')
        

